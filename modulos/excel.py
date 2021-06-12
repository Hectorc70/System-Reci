import os.path as path

import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook


class ArchivoExcel:
    """Clase Excel: lee o escribe en un archivo Excel"""

    def __init__(self, ruta):

        self.ruta = ruta.replace("/", "\\")
        self.archivo = self.ruta.split("\\")[-1]
        self.wb = self.__abrir_crear()
        self.hojas = self.__obtener_hojas()

        self.__str__()

    def __str__(self):

        return f"""Ruta del Archivo: {self.ruta}
					Nombre del Archivo: {self.archivo}"""

    def __obtener_hojas(self):
        """Retorna las hojas del archivo en un diccionario y una lista
                {'NOMBRE':NOMBRE_HOJA_TECNICA}, [NOMBRE_HOJA_TECNICA, 
                NOMBRE_HOJA_TECNICA2]"""

        hojas_nombres = {}

        for hoja_nombre, hojas in zip(self.wb.sheetnames, self.wb.worksheets):
            hojas_nombres[hoja_nombre] = hojas

        hojas_lista = [h for h in self.wb.worksheets]

        return [hojas_nombres, hojas_lista]

    def __abrir_crear(self):
        """"Abre o crear el archivo dado en el parametro 'ruta'"""

        if path.exists(self.ruta):

            documento = load_workbook(self.ruta)

        else:

            documento = Workbook()

        return documento

    def crear_hoja_nueva(self, nombre_hoja_nueva):
        """Crea una hoja nueva en el archivo excel"""

        self.hoja_nueva = self.wb.create_sheet(str(nombre_hoja_nueva))

    def leer_titulos(self, hoja, fila_de_lectura):
        """Obtiene titulos de columnas de la hoja 
                Pasada como parametro"""

        self.claves_columnas = dict()

        def obtener_claves_titulos(clave_titulo, texto_titulo):
            """Solo cuando se quiere leer"""

            celda = str(clave_titulo).split(".")[-1][:-2]
            titulo = texto_titulo[0]
            texto_depurado = str(titulo).strip(' ')
            clave_depurada = celda.strip('>')  # Clave de la columna

            return texto_depurado, clave_depurada

        if type(hoja) is str:

            hoja = self.hojas[0][hoja]
            titulos = hoja[fila_de_lectura]
            columna = 0
            for titulo in range(len(titulos)):

                if titulos[titulo].value is not None:

                    columna += 1

                    texto = [titulos[titulo].value]
                    clave = titulos[titulo]
                    claves = obtener_claves_titulos(clave, texto)
                    self.claves_columnas[claves[0]] = claves[1]

        elif type(hoja) is int:
            hoja = self.hojas[1][hoja]
            titulos = hoja[fila_de_lectura]

            for titulo in range(len(titulos)):
                if titulos[titulo].value is not None:

                    texto = [titulos[titulo].value]
                    clave = titulos[titulo]
                    claves = obtener_claves_titulos(clave, texto)

                    self.claves_columnas[claves[0]] = [claves[1]]

        return self.claves_columnas



    def escribir_titulo(self, titulos, fila, hoja_activa, col_ini=0, ):
        """Escribe los titulos en la fila y columna que se indique"""

        self.wb.active = hoja_activa

        if type(titulos) is list:
            for text_titulo in titulos:

                col_ini += 1
                titulo = self.wb.active.cell(row=fila, column=col_ini)
                titulo.value = (text_titulo)

        elif type(titulos) is str:
            titulo = self.wb.active.cell(row=fila, column=col_ini)
            titulo.value = (titulos)

    def escribir_en_hoja(self, contenido_lista, col, hoja_activa, fila_ini=1):
        """Escribe en la hoja que esta activa"""

        self.wb.active = hoja_activa

        if type(contenido_lista) is list and len(contenido_lista) == 1:
            for conte in contenido_lista:

                for texto in conte:

                    if type(texto) is list:
                        for text in texto:
                            fila_ini += 1
                            texto_celda = self.wb.active.cell(
                                row=fila_ini, column=col)
                            texto_celda.value = (text)
                    else:

                        fila_ini += 1
                        texto_celda = self.wb.active.cell(
                            row=fila_ini, column=col)
                        texto_celda.value = (texto)
        elif type(contenido_lista) is list and len(contenido_lista) >= 2:
            self.escribir_varias_columnas(
                hoja_activa, contenido_lista, col, fila_ini)

    def escribir_varias_columnas(self, hoja_activa, listas, col, fila):
        """SE DEBE PONER EL NUMERO DE LA COLUMNA QUE ESTA OCUPADA EJEMPLO
                O PARA QUE EMPIECE EN LA 1"""

        self.wb.active = hoja_activa

        for conte in listas:
            col += 1
            fila = 1

            if type(conte) is list:
                for texto in conte:
                    fila += 1

                    #print(texto,"|fila|",fila, "|columna|",col)

                    texto_celda = self.wb.active.cell(row=fila, column=col)
                    texto_celda.value = (texto)

                #print("leyendo Otra Lista")
            else:
                print("escritas todas las columnas")

    def guardar(self, ruta):

        self.wb.save(ruta)

        print(f"Guardado en la ruta: {ruta}")
